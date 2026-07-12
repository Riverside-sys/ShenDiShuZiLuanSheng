#!/usr/bin/env python3
"""从本地原始工作簿生成苏95单井岩性分层 TypeScript 数据。"""

from __future__ import annotations

import argparse
import json
import math
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPOSITORY_ROOT = Path(__file__).resolve().parents[2]
SOURCE_PATH = (
    REPOSITORY_ROOT
    / "src/Views/hanshuiceng/data/26-7吉大含水层资料/测井资料数字化/其它/苏95数据.xlsx"
)
OUTPUT_PATH = (
    REPOSITORY_ROOT / "src/data/aquifer/stratigraphy/su95.generated.ts"
)
REQUIRED_HEADERS = (
    "井号",
    "顶深",
    "底深",
    "层厚",
    "岩性",
    "_岩性_颜色",
    "_岩性_岩性",
    "_段",
)


def parse_arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--check",
        action="store_true",
        help="仅检查当前生成文件是否可由原始工作簿复现，不写入文件",
    )
    return parser.parse_args()


def require_text(value: Any, *, row_number: int, field: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise ValueError(f"第 {row_number} 行字段 {field!r} 不是非空文本")
    return value


def require_number(value: Any, *, row_number: int, field: str) -> int | float:
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"第 {row_number} 行字段 {field!r} 不是数值")
    if not math.isfinite(value):
        raise ValueError(f"第 {row_number} 行字段 {field!r} 不是有限数值")
    return value


def load_layers() -> tuple[str, list[str], list[dict[str, Any]]]:
    workbook = load_workbook(SOURCE_PATH, data_only=True, read_only=True)
    if len(workbook.sheetnames) != 1:
        raise ValueError(
            f"预期工作簿仅有 1 个 sheet，实际为 {workbook.sheetnames!r}"
        )

    worksheet = workbook[workbook.sheetnames[0]]
    header_cells = next(worksheet.iter_rows(min_row=1, max_row=1))
    headers = [cell.value for cell in header_cells]
    missing_headers = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing_headers:
        raise ValueError(f"工作簿缺少必需表头：{missing_headers!r}")

    layers: list[dict[str, Any]] = []
    for row_number, cells in enumerate(worksheet.iter_rows(min_row=2), start=2):
        values = [cell.value for cell in cells]
        if all(value is None for value in values):
            continue

        source = dict(zip(headers, values, strict=True))
        well_id = require_text(source["井号"], row_number=row_number, field="井号")
        if well_id != "苏95":
            raise ValueError(f"第 {row_number} 行井号为 {well_id!r}，不是“苏95”")

        top_depth = require_number(source["顶深"], row_number=row_number, field="顶深")
        bottom_depth = require_number(
            source["底深"], row_number=row_number, field="底深"
        )
        thickness = require_number(
            source["层厚"], row_number=row_number, field="层厚"
        )
        if top_depth >= bottom_depth:
            raise ValueError(
                f"第 {row_number} 行深度区间无效：{top_depth}–{bottom_depth}"
            )
        calculated_thickness = bottom_depth - top_depth
        if not math.isclose(thickness, calculated_thickness, abs_tol=1e-9):
            raise ValueError(
                f"第 {row_number} 行层厚 {thickness} 与深度差"
                f" {calculated_thickness} 不一致"
            )

        layers.append(
            {
                "sourceRow": row_number,
                "wellId": well_id,
                "topDepth": top_depth,
                "bottomDepth": bottom_depth,
                "thickness": thickness,
                "description": require_text(
                    source["岩性"], row_number=row_number, field="岩性"
                ),
                "color": require_text(
                    source["_岩性_颜色"],
                    row_number=row_number,
                    field="_岩性_颜色",
                ),
                "lithology": require_text(
                    source["_岩性_岩性"],
                    row_number=row_number,
                    field="_岩性_岩性",
                ),
                "sourceSectionCode": require_text(
                    source["_段"], row_number=row_number, field="_段"
                ),
            }
        )

    if not layers:
        raise ValueError("工作簿中没有分层记录")

    for current, following in zip(layers, layers[1:], strict=False):
        if current["topDepth"] >= following["topDepth"]:
            raise ValueError(
                f"第 {following['sourceRow']} 行未按顶深严格升序排列"
            )
        if current["bottomDepth"] > following["topDepth"]:
            raise ValueError(
                f"第 {current['sourceRow']} 与 {following['sourceRow']} 行深度重叠"
            )

    return worksheet.title, [str(header) for header in headers], layers


def render_typescript(layers: list[dict[str, Any]]) -> str:
    rendered_layers = ",\n".join(
        "    " + json.dumps(layer, ensure_ascii=False, separators=(",", ":"))
        for layer in layers
    )
    return f"""// 此文件由 scripts/aquifer/generate_su95_stratigraphy.py 生成，请勿手改。
// 数据源：本地原始资料“苏95数据.xlsx”的 Sheet1；原始 xlsx 不纳入 Git。
// “sourceSectionCode”仅保留工作簿“_段”列原值，不表示区域统一地层。
import type {{ Su95StratigraphyLayer }} from "./types.ts";

export const SU95_STRATIGRAPHY = Object.freeze([
{rendered_layers}
] satisfies readonly Su95StratigraphyLayer[]);
"""


def report(sheet_name: str, headers: list[str], layers: list[dict[str, Any]]) -> None:
    gaps = [
        following["topDepth"] - current["bottomDepth"]
        for current, following in zip(layers, layers[1:], strict=False)
        if following["topDepth"] > current["bottomDepth"]
    ]
    lithology_counts = Counter(layer["lithology"] for layer in layers)
    print(f"sheet: {sheet_name}")
    print(f"columns: {len(headers)}")
    print(f"headers: {headers}")
    print(f"records: {len(layers)}")
    print(
        f"depth range: {layers[0]['topDepth']}–{layers[-1]['bottomDepth']} m"
    )
    print(f"total thickness: {sum(layer['thickness'] for layer in layers)} m")
    print(f"depth gaps: {len(gaps)}, total {sum(gaps)} m")
    print(
        "lithology counts: "
        + json.dumps(dict(sorted(lithology_counts.items())), ensure_ascii=False)
    )


def main() -> None:
    arguments = parse_arguments()
    sheet_name, headers, layers = load_layers()
    generated_content = render_typescript(layers)
    report(sheet_name, headers, layers)

    if arguments.check:
        if not OUTPUT_PATH.exists():
            raise SystemExit(f"生成文件不存在：{OUTPUT_PATH}")
        if OUTPUT_PATH.read_text(encoding="utf-8") != generated_content:
            raise SystemExit("生成文件与原始工作簿不一致，请重新运行生成脚本")
        print(f"reproducibility check passed: {OUTPUT_PATH.relative_to(REPOSITORY_ROOT)}")
        return

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(generated_content, encoding="utf-8")
    print(f"written: {OUTPUT_PATH.relative_to(REPOSITORY_ROOT)}")


if __name__ == "__main__":
    main()
