#!/usr/bin/env python3
"""Generate deterministic, traceable aquifer log data from local XLSX sources."""

from __future__ import annotations

import json
import math
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REPOSITORY_ROOT = Path(__file__).resolve().parents[2]
SOURCE_DIRECTORY = (
    REPOSITORY_ROOT
    / "src/Views/hanshuiceng/data/26-7吉大含水层资料/测井资料数字化"
)
OUTPUT_FILE = REPOSITORY_ROOT / "src/data/aquifer/logs/logs.generated.ts"
MAX_VISUALIZATION_SAMPLES = 1_200


@dataclass(frozen=True)
class ChannelSpec:
    """Maps an exact workbook column label to a stable application channel ID."""

    channel_id: str
    source_label: str
    interpretation: str


@dataclass(frozen=True)
class WorkbookSpec:
    """Describes the workbook structure that must be present for generation."""

    filename: str
    well_id: str
    sheet_names: tuple[str, ...]
    channels: tuple[ChannelSpec, ...]
    well_id_mapping_note: str | None = None


SP = ChannelSpec("sp", "自然电位", "自然电位（SP）；名称来自原表")
RESISTIVITY = ChannelSpec("resistivity", "电阻率", "电阻率；名称来自原表")

WORKBOOK_SPECS = (
    WorkbookSpec(
        "N参1-声波补偿2800开始.xlsx",
        "N参1",
        ("Sheet1",),
        (
            ChannelSpec(
                "naturalEncoded",
                "自然加码",
                "原表列名为“自然加码”，不推断其物理含义",
            ),
            SP,
            ChannelSpec(
                "compensatedDensity",
                "补偿密度",
                "补偿密度；名称来自原表",
            ),
            ChannelSpec(
                "compensatedAcoustic",
                "补偿声波",
                "补偿声波；名称来自原表",
            ),
        ),
    ),
    WorkbookSpec(
        "华洋3.xlsx",
        "洋3",
        ("Sheet1",),
        (
            SP,
            ChannelSpec(
                "resistivity",
                "2.5米电阻率",
                "2.5米电阻率；名称来自原表",
            ),
        ),
        "工作簿文件名为“华洋3”，井名称单元格为“洋3”",
    ),
    WorkbookSpec(
        "新3.xlsx",
        "新3",
        ("Sheet1", "Sheet2"),
        (
            SP,
            ChannelSpec(
                "resistivity",
                "2.5米电阻",
                "2.5米电阻；按项目曲线用途归入 resistivity，原名完整保留",
            ),
        ),
    ),
    WorkbookSpec(
        "涟1.xlsx",
        "涟1",
        ("Sheet1", "Sheet2"),
        (
            SP,
            ChannelSpec(
                "resistivity",
                "电阻率（2.5米底）",
                "电阻率（2.5米底）；名称来自原表",
            ),
        ),
    ),
    WorkbookSpec("苏107.xlsx", "苏107", ("Sheet1", "Sheet2"), (SP, RESISTIVITY)),
    WorkbookSpec("苏118.xlsx", "苏118", ("Sheet1", "Sheet2"), (SP, RESISTIVITY)),
    WorkbookSpec("苏80.xlsx", "苏80", ("Sheet1", "Sheet2"), (SP, RESISTIVITY)),
    WorkbookSpec("阜3.xlsx", "阜3", ("Sheet1", "Sheet2"), (SP, RESISTIVITY)),
)


def require_finite_number(value: Any, context: str) -> float:
    """Returns a finite float or raises with source context."""

    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"{context}: expected a number, got {value!r}")
    number = float(value)
    if not math.isfinite(number):
        raise ValueError(f"{context}: expected a finite number, got {value!r}")
    return number


def optional_finite_number(value: Any, context: str) -> float | None:
    """Returns None for an empty cell and rejects non-finite source values."""

    if value is None:
        return None
    return require_finite_number(value, context)


def compact_number(value: float) -> int | float:
    """Emits integral floats as integers without changing their value."""

    return int(value) if value.is_integer() else value


def find_header_row(worksheet: Any) -> int:
    """Finds the one-based row containing the exact depth header."""

    matching_rows = [
        row_number
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True), 1)
        if row and row[0] == "深度"
    ]
    if len(matching_rows) != 1:
        raise ValueError(
            f"{worksheet.title}: expected one 深度 header, found {matching_rows}"
        )
    return matching_rows[0]


def read_sheet_structure(worksheet: Any) -> dict[str, Any]:
    """Captures sheet dimensions, exact headers, and numeric depth-row count."""

    header_row = find_header_row(worksheet)
    headers = [
        "" if value is None else str(value).strip()
        for value in next(
            worksheet.iter_rows(
                min_row=header_row,
                max_row=header_row,
                values_only=True,
            )
        )
    ]
    data_row_count = sum(
        1
        for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True)
        if row
        and isinstance(row[0], (int, float))
        and not isinstance(row[0], bool)
        and math.isfinite(float(row[0]))
    )
    return {
        "name": worksheet.title,
        "rowCount": worksheet.max_row,
        "columnCount": worksheet.max_column,
        "headerRow": header_row,
        "headers": headers,
        "dataRowCount": data_row_count,
    }


def read_declared_number(worksheet: Any, label: str, header_row: int) -> float:
    """Reads a required numeric workbook declaration above the data header."""

    for row in worksheet.iter_rows(
        min_row=1,
        max_row=header_row - 1,
        values_only=True,
    ):
        if row and row[0] == label:
            value = row[1] if len(row) > 1 else None
            return require_finite_number(value, f"{worksheet.title} {label}")
    raise ValueError(f"{worksheet.title}: missing declaration {label!r}")


def read_source_rows(
    worksheet: Any,
    spec: WorkbookSpec,
    header_row: int,
) -> list[tuple[float, list[float | None]]]:
    """Reads depth and configured channels without filling missing cells."""

    expected_headers = ["深度", *(channel.source_label for channel in spec.channels)]
    actual_headers = [
        worksheet.cell(header_row, column).value
        for column in range(1, len(expected_headers) + 1)
    ]
    if actual_headers != expected_headers:
        raise ValueError(
            f"{spec.filename} {worksheet.title}: headers changed; "
            f"expected {expected_headers!r}, got {actual_headers!r}"
        )

    rows: list[tuple[float, list[float | None]]] = []
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
        header_row + 1,
    ):
        if not row or all(value is None for value in row):
            continue
        depth = require_finite_number(
            row[0],
            f"{spec.filename} {worksheet.title} row {row_number} depth",
        )
        values = [
            optional_finite_number(
                row[column],
                (
                    f"{spec.filename} {worksheet.title} row {row_number} "
                    f"{channel.source_label}"
                ),
            )
            for column, channel in enumerate(spec.channels, 1)
        ]
        rows.append((depth, values))

    if len(rows) < 2:
        raise ValueError(f"{spec.filename}: expected at least two source rows")
    if any(current[0] <= previous[0] for previous, current in zip(rows, rows[1:])):
        raise ValueError(f"{spec.filename}: source depth must be strictly increasing")
    return rows


def validate_mirror_sheet(
    workbook: Any,
    spec: WorkbookSpec,
    source_rows: list[tuple[float, list[float | None]]],
) -> None:
    """Verifies Sheet2 mirrors Sheet1 depth/channels when the workbook has it."""

    if "Sheet2" not in spec.sheet_names:
        return

    worksheet = workbook["Sheet2"]
    header_row = find_header_row(worksheet)
    expected_headers = [
        "深度",
        "埋深",
        *(channel.source_label for channel in spec.channels),
    ]
    actual_headers = [
        worksheet.cell(header_row, column).value
        for column in range(1, len(expected_headers) + 1)
    ]
    if actual_headers != expected_headers:
        raise ValueError(
            f"{spec.filename} Sheet2: headers changed; "
            f"expected {expected_headers!r}, got {actual_headers!r}"
        )

    mirror_rows: list[tuple[float, list[float | None]]] = []
    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
        header_row + 1,
    ):
        if not row or all(value is None for value in row):
            continue
        depth = require_finite_number(
            row[0],
            f"{spec.filename} Sheet2 row {row_number} depth",
        )
        values = [
            optional_finite_number(
                row[column],
                (
                    f"{spec.filename} Sheet2 row {row_number} "
                    f"{channel.source_label}"
                ),
            )
            for column, channel in enumerate(spec.channels, 2)
        ]
        mirror_rows.append((depth, values))

    if mirror_rows != source_rows:
        raise ValueError(f"{spec.filename}: Sheet2 no longer mirrors Sheet1")


def channel_statistics(
    rows: list[tuple[float, list[float | None]]],
    channel_index: int,
    spec: ChannelSpec,
) -> dict[str, Any]:
    """Computes counts and observed range over unmodified source values."""

    values = [
        row_values[channel_index]
        for _, row_values in rows
        if row_values[channel_index] is not None
    ]
    if not values:
        raise ValueError(f"{spec.source_label}: channel has no finite values")
    finite_values = [value for value in values if value is not None]
    return {
        "id": spec.channel_id,
        "sourceLabel": spec.source_label,
        "interpretation": spec.interpretation,
        "unit": None,
        "validSampleCount": len(finite_values),
        "missingSampleCount": len(rows) - len(finite_values),
        "minimum": compact_number(min(finite_values)),
        "maximum": compact_number(max(finite_values)),
    }


def nearest_valid_neighbors(
    rows: list[tuple[float, list[float | None]]],
    channel_index: int,
) -> tuple[list[int | None], list[int | None]]:
    """Finds nearest valid source values before and after every row."""

    previous: list[int | None] = [None] * len(rows)
    next_: list[int | None] = [None] * len(rows)

    last_valid: int | None = None
    for index, (_, values) in enumerate(rows):
        previous[index] = last_valid
        if values[channel_index] is not None:
            last_valid = index

    last_valid = None
    for index in range(len(rows) - 1, -1, -1):
        next_[index] = last_valid
        if rows[index][1][channel_index] is not None:
            last_valid = index
    return previous, next_


def local_importance_scores(
    rows: list[tuple[float, list[float | None]]],
    statistics: list[dict[str, Any]],
) -> list[float]:
    """Scores multichannel deviation from neighboring measured-value trends."""

    scores = [0.0] * len(rows)
    for channel_index, channel in enumerate(statistics):
        value_range = float(channel["maximum"]) - float(channel["minimum"])
        if value_range == 0:
            continue
        previous, next_ = nearest_valid_neighbors(rows, channel_index)
        for index, (depth, values) in enumerate(rows):
            value = values[channel_index]
            left_index = previous[index]
            right_index = next_[index]
            if value is None or left_index is None or right_index is None:
                continue
            left_depth, left_values = rows[left_index]
            right_depth, right_values = rows[right_index]
            left_value = left_values[channel_index]
            right_value = right_values[channel_index]
            if left_value is None or right_value is None or right_depth == left_depth:
                continue
            ratio = (depth - left_depth) / (right_depth - left_depth)
            expected = left_value + (right_value - left_value) * ratio
            scores[index] = max(scores[index], abs(value - expected) / value_range)
    return scores


def mandatory_sample_indices(
    rows: list[tuple[float, list[float | None]]],
) -> set[int]:
    """Keeps endpoints plus each channel's measured endpoints and global extrema."""

    indices = {0, len(rows) - 1}
    channel_count = len(rows[0][1])
    for channel_index in range(channel_count):
        measured = [
            (index, values[channel_index])
            for index, (_, values) in enumerate(rows)
            if values[channel_index] is not None
        ]
        if not measured:
            continue
        indices.update((measured[0][0], measured[-1][0]))
        indices.add(min(measured, key=lambda item: (float(item[1]), item[0]))[0])
        indices.add(max(measured, key=lambda item: (float(item[1]), -item[0]))[0])
    return indices


def sample_indices(
    rows: list[tuple[float, list[float | None]]],
    statistics: list[dict[str, Any]],
    limit: int,
) -> list[int]:
    """Selects deterministic trend/extrema samples without interpolation."""

    if len(rows) <= limit:
        return list(range(len(rows)))

    selected = mandatory_sample_indices(rows)
    available_slots = limit - len(selected)
    bucket_count = max(1, available_slots // 2)
    scores = local_importance_scores(rows, statistics)
    interior_count = len(rows) - 2

    for bucket_index in range(bucket_count):
        start = 1 + (bucket_index * interior_count) // bucket_count
        end = 1 + ((bucket_index + 1) * interior_count) // bucket_count
        if start >= end:
            continue
        trend_index = (start + end - 1) // 2
        extrema_index = max(range(start, end), key=lambda index: (scores[index], -index))
        selected.add(trend_index)
        selected.add(extrema_index)

    if len(selected) < limit:
        for slot in range(limit):
            index = round(slot * (len(rows) - 1) / (limit - 1))
            selected.add(index)
            if len(selected) == limit:
                break

    if len(selected) < limit:
        for index in range(len(rows)):
            selected.add(index)
            if len(selected) == limit:
                break

    if len(selected) > limit:
        mandatory = mandatory_sample_indices(rows)
        removable = sorted(
            selected - mandatory,
            key=lambda index: (scores[index], index),
        )
        for index in removable[: len(selected) - limit]:
            selected.remove(index)

    return sorted(selected)


def generate_well(spec: WorkbookSpec) -> dict[str, Any]:
    """Loads, validates, summarizes, and samples one source workbook."""

    workbook_path = SOURCE_DIRECTORY / spec.filename
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Missing source workbook: {workbook_path}")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if tuple(workbook.sheetnames) != spec.sheet_names:
            raise ValueError(
                f"{spec.filename}: expected sheets {spec.sheet_names!r}, "
                f"got {tuple(workbook.sheetnames)!r}"
            )

        sheet_structures = [
            read_sheet_structure(workbook[sheet_name])
            for sheet_name in spec.sheet_names
        ]
        source_sheet = workbook["Sheet1"]
        header_row = find_header_row(source_sheet)
        rows = read_source_rows(source_sheet, spec, header_row)
        validate_mirror_sheet(workbook, spec, rows)

        statistics = [
            channel_statistics(rows, channel_index, channel)
            for channel_index, channel in enumerate(spec.channels)
        ]
        selected_indices = sample_indices(
            rows,
            statistics,
            MAX_VISUALIZATION_SAMPLES,
        )
        samples = [
            [
                compact_number(rows[index][0]),
                *[
                    None if value is None else compact_number(value)
                    for value in rows[index][1]
                ],
            ]
            for index in selected_indices
        ]

        source: dict[str, Any] = {
            "workbookName": spec.filename,
            "selectedSheetName": source_sheet.title,
            "headerRow": header_row,
            "sheets": sheet_structures,
        }
        if spec.well_id_mapping_note:
            source["wellIdMappingNote"] = spec.well_id_mapping_note

        return {
            "wellId": spec.well_id,
            "source": source,
            "declaredDepthRange": {
                "minimum": compact_number(
                    read_declared_number(source_sheet, "初始深度:", header_row)
                ),
                "maximum": compact_number(
                    read_declared_number(source_sheet, "结束深度:", header_row)
                ),
            },
            "sourceDepthRange": {
                "minimum": compact_number(rows[0][0]),
                "maximum": compact_number(rows[-1][0]),
            },
            "declaredSamplingInterval": compact_number(
                read_declared_number(source_sheet, "采样间隔:", header_row)
            ),
            "originalSampleCount": len(rows),
            "isVisualizationSampled": len(selected_indices) < len(rows),
            "channels": statistics,
            "channelOrder": [channel.channel_id for channel in spec.channels],
            "samples": samples,
        }
    finally:
        workbook.close()


def format_generated_well(well: dict[str, Any]) -> str:
    """Formats sample tuples compactly while keeping metadata reviewable."""

    placeholder = "__SAMPLE_ROWS__"
    serializable = {**well, "samples": placeholder}
    formatted = json.dumps(
        serializable,
        ensure_ascii=False,
        indent=2,
        allow_nan=False,
    )
    sample_rows = ",\n".join(
        f"    {json.dumps(row, ensure_ascii=False, separators=(',', ':'))}"
        for row in well["samples"]
    )
    formatted_samples = f"[\n{sample_rows}\n  ]"
    return formatted.replace(f'"{placeholder}"', formatted_samples)


def render_typescript(wells: list[dict[str, Any]]) -> str:
    """Renders the generated data as a dependency-free TypeScript module."""

    entries = ",\n".join(
        "\n".join(f"  {line}" for line in format_generated_well(well).splitlines())
        for well in wells
    )
    return f"""// Generated by scripts/aquifer/generate_logs.py. Do not edit manually.
// Values are source workbook cells; null means the original cell was empty.

export interface GeneratedAquiferLog {{
    wellId: string;
    source: {{
        workbookName: string;
        selectedSheetName: string;
        headerRow: number;
        wellIdMappingNote?: string;
        sheets: Array<{{
            name: string;
            rowCount: number;
            columnCount: number;
            headerRow: number;
            headers: string[];
            dataRowCount: number;
        }}>;
    }};
    declaredDepthRange: {{ minimum: number; maximum: number }};
    sourceDepthRange: {{ minimum: number; maximum: number }};
    declaredSamplingInterval: number;
    originalSampleCount: number;
    isVisualizationSampled: boolean;
    channels: Array<{{
        id: string;
        sourceLabel: string;
        interpretation: string;
        unit: null;
        validSampleCount: number;
        missingSampleCount: number;
        minimum: number;
        maximum: number;
    }}>;
    channelOrder: string[];
    samples: Array<[number, ...(number | null)[]]>;
}}

export const GENERATED_AQUIFER_LOGS: GeneratedAquiferLog[] = [
{entries}
];
"""


def main() -> None:
    """Generates the TypeScript dataset and reports deterministic counts."""

    wells = [generate_well(spec) for spec in WORKBOOK_SPECS]
    output = render_typescript(wells)
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_FILE.write_text(output, encoding="utf-8")

    original_count = sum(well["originalSampleCount"] for well in wells)
    sample_count = sum(len(well["samples"]) for well in wells)
    print(
        f"Generated {OUTPUT_FILE.relative_to(REPOSITORY_ROOT)}: "
        f"{len(wells)} wells, {original_count} source rows, "
        f"{sample_count} visualization samples"
    )


if __name__ == "__main__":
    main()
